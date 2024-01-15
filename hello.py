from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from bs4 import BeautifulSoup
import time
import pandas as pd  
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Set path to the geckodriver
gecko_driver_path = r'C:\Users\ee\Desktop\geckodriver-v0.34.0-win64\geckodriver.exe'

# Initialize the WebDriver with Firefox
service = Service(gecko_driver_path)
driver = webdriver.Firefox(service=service)

url = 'https://www.nike.com/w/mens-jordan-shoes-37eefznik1zy7ok'
driver.get(url)

# Gradual scrolling
scroll_increment = 2000
last_height = driver.execute_script("return document.body.scrollHeight")
retries = 0
max_retries = 6  # Maximum number of retries at the bottom of the page

while True:
    driver.execute_script(f"window.scrollBy(0, {scroll_increment});")
    time.sleep(3)  # Increased sleep time

    new_height = driver.execute_script("return document.body.scrollHeight")
    if new_height == last_height:
        retries += 1
        if retries >= max_retries:
            break
    else:
        retries = 0
    last_height = new_height

# Now that the page is fully loaded, get the page source
page_source = driver.page_source
soup = BeautifulSoup(page_source, 'html.parser')

products_data = []


# Extract and print product names
for product in soup.find_all('div', class_='product-card__body'):
    name = product.find('div', class_='product-card__title').text.strip() if product.find('div', class_='product-card__title') else 'No Name'
    price = product.find('div', class_='product-price').text.strip() if product.find('div', class_='product-price') else 'No Price'
    promo = product.find('div', class_='at-pw-promo').text.strip() if product.find('div', class_='at-pw-promo') else 'No Promo'

    print(f"{name}: {price}  {promo}")
    products_data.append({'Name': name, 'Price': price, 'Promo': promo})


# Close the browser
driver.quit()

# File path for the Excel file
file_path = 'nike_products.xlsx'

# Check if the file exists
if os.path.exists(file_path):
    # Read the existing data
    existing_data = pd.read_excel(file_path, engine='openpyxl')
    # Convert new data to DataFrame
    new_data = pd.DataFrame(products_data)
    # Concatenate new data
    updated_data = pd.concat([existing_data, new_data], ignore_index=True)
else:
    # Convert new data to DataFrame
    updated_data = pd.DataFrame(products_data)

# Create a Pandas Excel writer using openpyxl as the engine
writer = pd.ExcelWriter(file_path, engine='openpyxl')

# Write the DataFrame data to the Excel file
updated_data.to_excel(writer, index=False, sheet_name='Sheet1')

# Access the openpyxl workbook and worksheet
workbook = writer.book
if not workbook.worksheets:
    workbook.create_sheet('Sheet1')
worksheet = workbook['Sheet1']

# Set the column widths based on the max length in each column
for column in worksheet.columns:
    max_length = 0
    column = [cell for cell in column if cell.value]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Close the Pandas Excel writer and save the Excel file
writer.close()


